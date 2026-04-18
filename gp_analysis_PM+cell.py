"""
GP Value Analysis Pipeline for Fluorescence Microscopy (CZI files)
===================================================================
Calculates Generalized Polarization (GP) values for membrane-stained cells.

GP = (I_440 - I_490_mean) / (I_440 + I_490_mean)
where I_490_mean = mean of two 490nm measurements (bleaching correction)

Channel layout (auto-detected, adjust CH_* constants if needed):
  C0: Brightfield (transmitted)
  C1: 490nm measurement A
  C2: 440nm (also used for segmentation)
  C3: 490nm measurement B
  C4: 650nm Cy5 (SpyDNA, not used in GP)

Usage:
  Place this script in the folder that contains your CZI images and run:
      python gp_analysis.py

Outputs (saved in ./results/ subfolder):
  - results/gp_results.xlsx                  : all cells, all images, one sheet
  - THRESHOLD_FACTOR      : fraction of Otsu threshold (1.0 = bright compact cells only,
                             0.3 = also catches dim flat adhered cells — default)
  - MIN_CELL_AREA / MAX_CELL_AREA : size filter in px²
  - results/<image>/<image>_envelope.tif     : labeled membrane-ring mask
  - results/<image>/<image>_overview.png     : annotated overview image
"""

import os
import sys
import glob
import struct
import warnings
import numpy as np
import pandas as pd
from pathlib import Path
from scipy import ndimage
from skimage import filters, morphology, measure, segmentation
from skimage.feature import peak_local_max
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from tifffile import imwrite
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURABLE PARAMETERS  ← adjust these to tune the analysis
# ─────────────────────────────────────────────────────────────────────────────

# Channel indices within the CZI file (0-based)
CH_BRIGHTFIELD = 0
CH_490_A       = 1   # first 490nm measurement
CH_440         = 2   # 440nm  (also used for segmentation)
CH_490_B       = 3   # second 490nm measurement (bleaching correction)
CH_650         = 4   # 650nm Cy5 / SpyDNA

# Microscope calibration
PIXEL_SIZE_UM         = 0.11  # µm per pixel — FALLBACK ONLY (used if metadata extraction fails)
                               # The script automatically reads pixel size from each CZI file's metadata.
                               # This value is only used if that extraction fails (rare).
                               # Typical values: 0.05-0.07 µm/px (1×1 binning), 0.10-0.14 µm/px (2×2 binning)

# Segmentation parameters
BG_CORRECTION_SIGMA   = 80    # sigma (px) for large-scale background estimation
SEGMENTATION_SIGMA    = 2     # Gaussian blur before thresholding
THRESHOLD_PERCENTILE  = 92    # image-adaptive threshold: pixels above this percentile
                               # of the bg-corrected image are considered foreground.
                               # 92 works well for fields with sparse cells (few bright px).
                               # Lower (e.g. 88) if dim flat cells are missed;
                               # raise (e.g. 95) if too much background noise is included.
CLOSING_RADIUS        = 8     # morphological closing radius (px) applied after thresholding
                               # — bridges small gaps in dim flat cell membranes
MIN_CELL_DIAMETER_UM  = 8.0  # µm — cells smaller than this diameter are rejected as debris
MAX_CELL_DIAMETER_UM  = 40.0 # µm — cells larger than this are rejected as aggregates
MIN_CIRCULARITY       = 0.15  # 0–1; very low to allow irregular flat spread-out cells
WATERSHED_MIN_DIST    = 20    # minimum distance (px) between watershed seeds

# Membrane ring
MEMBRANE_THICKNESS_UM = 0.8   # µm — thickness of membrane ring used for GP measurement
                               # The script converts this to pixels based on each image's calibration.
                               # Typical plasma membrane: 0.5-1.0 µm thick ring captures membrane signal

# ─────────────────────────────────────────────────────────────────────────────


# ── CZI reader ────────────────────────────────────────────────────────────────

def find_czi_subblocks(filepath):
    """Return list of (position, channel_index) for all SubBlock segments."""
    subblocks = []
    with open(filepath, 'rb') as f:
        while True:
            pos = f.tell()
            header = f.read(32)
            if len(header) < 32:
                break
            sid = header[:16].rstrip(b'\x00').decode('ascii', errors='ignore')
            alloc = struct.unpack('<q', header[16:24])[0]
            if sid == 'ZISRAWSUBBLOCK':
                # Read DirectoryEntryDV to find channel
                f.seek(pos + 32 + 16)  # skip 16 bytes of subblock preamble header info
                # Actually: subblock header layout:
                #   4 bytes metadata_size
                #   4 bytes attach_size  (deprecated, always 0)
                #   8 bytes data_size
                # then DirectoryEntryDV starts
                f.seek(pos + 32)
                meta_size = struct.unpack('<i', f.read(4))[0]
                f.read(4)   # attachsize
                f.read(8)   # datasize
                # DirectoryEntryDV:
                f.read(2)   # schema
                f.read(4)   # pixeltype
                f.read(8)   # filepos
                f.read(4)   # filepart
                f.read(4)   # compression
                f.read(1)   # pyramid_type
                f.read(1)   # spare
                f.read(4)   # spare
                ndim = struct.unpack('<i', f.read(4))[0]
                channel = 0
                for _ in range(ndim):
                    dim = f.read(4).decode('ascii', errors='ignore').rstrip('\x00')
                    start = struct.unpack('<i', f.read(4))[0]
                    f.read(4 + 4 + 4)  # size, float_size, stored_size
                    if dim == 'C':
                        channel = start
                subblocks.append((pos, channel, meta_size))
            if alloc > 0:
                f.seek(pos + 32 + alloc)
            else:
                break
    return sorted(subblocks, key=lambda x: x[1])  # sort by channel


def read_channel(filepath, pos, meta_size, height=1028, width=1216):
    """Read a single channel image from a CZI SubBlock."""
    with open(filepath, 'rb') as f:
        data_offset = pos + 32 + 256 + meta_size
        f.seek(data_offset)
        raw = f.read(height * width * 2)
    return np.frombuffer(raw, dtype=np.uint16).reshape(height, width).copy()


def load_czi_channels(filepath):
    """Load all channels from a CZI file. Returns dict {channel_idx: array}, height, width, pixel_size_um."""
    subblocks = find_czi_subblocks(filepath)
    if not subblocks:
        raise ValueError(f"No image data found in {filepath}")

    # Determine image dimensions from first block
    pos0, _, meta0 = subblocks[0]
    # Quick peek at the directory entry to get X/Y size
    with open(filepath, 'rb') as f:
        f.seek(pos0 + 32)
        meta_size_h = struct.unpack('<i', f.read(4))[0]
        f.read(4); f.read(8); f.read(2); f.read(4); f.read(8); f.read(4)
        f.read(4); f.read(1); f.read(1); f.read(4)
        ndim = struct.unpack('<i', f.read(4))[0]
        dims = {}
        for _ in range(ndim):
            dim = f.read(4).decode('ascii', errors='ignore').rstrip('\x00')
            start = struct.unpack('<i', f.read(4))[0]
            size  = struct.unpack('<i', f.read(4))[0]
            f.read(4); f.read(4)
            dims[dim] = size
    height = dims.get('Y', 1028)
    width  = dims.get('X', 1216)

    # Extract pixel size from metadata
    pixel_size_um = _extract_pixel_size_from_metadata(filepath)

    channels = {}
    for pos, ch_idx, meta_size in subblocks:
        channels[ch_idx] = read_channel(filepath, pos, meta_size, height, width)
    return channels, height, width, pixel_size_um


def _extract_pixel_size_from_metadata(filepath):
    """Extract pixel size (µm/px) from CZI metadata XML."""
    with open(filepath, 'rb') as f:
        while True:
            pos = f.tell()
            header = f.read(32)
            if len(header) < 32:
                break
            sid = header[:16].rstrip(b'\x00').decode('ascii', errors='ignore')
            alloc = struct.unpack('<q', header[16:24])[0]
            if sid == 'ZISRAWMETADATA':
                f.seek(pos + 32)
                meta_header = f.read(256)
                xml_size = struct.unpack('<i', meta_header[:4])[0]
                f.seek(pos + 32 + 256)
                xml_data = f.read(xml_size)
                xml_str = xml_data.decode('utf-8', errors='ignore')
                
                # Look for <Distance Id="X"><Value>NUM</Value> (in meters)
                import re
                x_match = re.search(r'<Distance[^>]*Id="X"[^>]*>.*?<Value>([\d.eE+-]+)</Value>',
                                    xml_str, re.DOTALL | re.IGNORECASE)
                if x_match:
                    pixel_size_m = float(x_match.group(1))
                    return pixel_size_m * 1e6  # convert to µm
                break
            if alloc > 0:
                f.seek(pos + 32 + alloc)
            else:
                break
    # Fallback if not found
    print(f"  WARNING: Could not extract pixel size from metadata, using default {PIXEL_SIZE_UM} µm/px")
    return PIXEL_SIZE_UM


# ── Segmentation ──────────────────────────────────────────────────────────────

def segment_cells(ch440, min_diameter_um, max_diameter_um, min_circularity,
                  pixel_size_um, bg_sigma, seg_sigma, threshold_percentile,
                  closing_radius, watershed_min_dist):
    """
    Robust segmentation for mixed populations of bright compact cells AND
    dim flat spread-out adhered cells.

    Strategy:
      1. Background-correct with large Gaussian to remove illumination gradient
      2. Threshold at a fixed image percentile (adaptive, not Otsu-based)
         — percentile threshold is robust regardless of how bright or dim cells are
      3. Morphological closing to bridge gaps in dim flat cell membranes
      4. Hole-filling, debris removal
      5. Watershed to separate touching cells
      6. Size (diameter) + circularity filter

    Args:
        min_diameter_um, max_diameter_um: physical size filter in µm
        pixel_size_um: microscope calibration (µm per pixel)

    Returns:
        labeled_cells : 2D int array, 0=background, N=cell N
        n_cells       : number of accepted cells
    """
    img = ch440.astype(np.float32)

    # Convert diameter limits to area in pixels
    min_area_um2 = np.pi * (min_diameter_um / 2) ** 2
    max_area_um2 = np.pi * (max_diameter_um / 2) ** 2
    min_area_px  = min_area_um2 / (pixel_size_um ** 2)
    max_area_px  = max_area_um2 / (pixel_size_um ** 2)

    # 1. Background correction
    bg      = filters.gaussian(img, sigma=bg_sigma)
    corr    = img - bg

    # 2. Smooth and threshold at fixed percentile (image-adaptive)
    blurred = filters.gaussian(corr, sigma=seg_sigma)
    thresh  = np.percentile(blurred, threshold_percentile)
    binary  = blurred > thresh

    # 3. Morphological closing to connect fragmented flat cell signal
    closed  = morphology.closing(binary, morphology.disk(closing_radius))

    # 4. Fill holes (handles hollow-looking flat cells)
    filled  = ndimage.binary_fill_holes(closed)

    # 5. Remove small debris
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", FutureWarning)
        cleaned = morphology.remove_small_objects(filled, min_size=int(min_area_px) + 1)

    # 6. Watershed to separate touching cells
    distance   = ndimage.distance_transform_edt(cleaned)
    coords     = peak_local_max(distance, min_distance=watershed_min_dist, labels=cleaned)
    mask_peaks = np.zeros(distance.shape, dtype=bool)
    mask_peaks[tuple(coords.T)] = True
    markers, _ = ndimage.label(mask_peaks)
    labeled_ws = segmentation.watershed(-distance, markers, mask=cleaned)

    # 7. Filter by diameter (in µm) and circularity; renumber sequentially
    labeled_cells = np.zeros_like(labeled_ws, dtype=np.int32)
    new_label = 1
    rejected  = []
    for prop in measure.regionprops(labeled_ws):
        area  = prop.area
        perim = prop.perimeter if prop.perimeter > 0 else 1
        circ  = 4 * np.pi * area / (perim ** 2)
        # Convert area to diameter in µm
        diameter_um = 2 * np.sqrt(area / np.pi) * pixel_size_um
        
        if min_area_px <= area <= max_area_px and circ >= min_circularity:
            labeled_cells[labeled_ws == prop.label] = new_label
            new_label += 1
        else:
            rejected.append((prop.label, round(diameter_um, 1), round(circ, 2)))

    if rejected:
        print(f"    Rejected {len(rejected)} object(s):")
        for lbl, diam, c in rejected[:10]:  # show first 10
            print(f"      #{lbl}: diameter={diam}µm, circ={c}")
        if len(rejected) > 10:
            print(f"      ... and {len(rejected)-10} more")

    return labeled_cells, new_label - 1


# ── Membrane ring extraction ──────────────────────────────────────────────────

def make_membrane_masks(labeled_cells, thickness_um, pixel_size_um):
    """
    For each cell in labeled_cells, erode inward by `thickness_um` (in µm)
    and return the ring region as a labeled array (same cell numbers).
    
    Args:
        thickness_um: membrane ring thickness in micrometers
        pixel_size_um: pixel size calibration (µm per pixel)
    """
    # Convert thickness from µm to pixels
    thickness_px = int(round(thickness_um / pixel_size_um))
    if thickness_px < 1:
        thickness_px = 1
    
    membrane = np.zeros_like(labeled_cells, dtype=np.int32)
    disk     = morphology.disk(thickness_px)
    for cell_id in np.unique(labeled_cells):
        if cell_id == 0:
            continue
        cell_mask = labeled_cells == cell_id
        eroded    = morphology.erosion(cell_mask, disk)
        ring      = cell_mask & ~eroded
        membrane[ring] = cell_id
    return membrane


# ── Intensity measurements ────────────────────────────────────────────────────

def measure_intensities(labeled_cells, membrane_mask, ch_440, ch_490a, ch_490b):
    """
    Measure per-pixel mean intensities for background, each cell's membrane, AND whole cell.

    Returns a list of dicts (one per cell) with both membrane and whole-cell measurements.
    Background = mean of all pixels NOT belonging to any cell.
    """
    bg_mask = labeled_cells == 0

    results = []
    bg_440  = float(np.mean(ch_440[bg_mask]))  if bg_mask.any() else np.nan
    bg_490a = float(np.mean(ch_490a[bg_mask])) if bg_mask.any() else np.nan
    bg_490b = float(np.mean(ch_490b[bg_mask])) if bg_mask.any() else np.nan

    for cell_id in sorted(np.unique(labeled_cells)):
        if cell_id == 0:
            continue
            
        # MEMBRANE measurements
        ring = membrane_mask == cell_id
        if not ring.any():
            print(f"    Cell {cell_id}: ring empty after erosion (cell too small), skipping")
            continue
        MIN_RING_PIXELS = 50
        if ring.sum() < MIN_RING_PIXELS:
            print(f"    Cell {cell_id}: ring only {ring.sum()} px (< {MIN_RING_PIXELS}), skipping")
            continue

        # Raw membrane intensities (mean per pixel)
        raw_memb_440  = float(np.mean(ch_440[ring]))
        raw_memb_490a = float(np.mean(ch_490a[ring]))
        raw_memb_490b = float(np.mean(ch_490b[ring]))

        # Background-subtracted membrane
        sub_memb_440  = raw_memb_440  - bg_440
        sub_memb_490a = raw_memb_490a - bg_490a
        sub_memb_490b = raw_memb_490b - bg_490b

        # GP for membrane
        I_490_memb = (sub_memb_490a + sub_memb_490b) / 2.0
        I_440_memb = sub_memb_440
        denom_memb = I_440_memb + I_490_memb
        gp_memb = (I_440_memb - I_490_memb) / denom_memb if denom_memb != 0 else np.nan

        # WHOLE CELL measurements
        whole_cell = labeled_cells == cell_id
        
        # Raw whole-cell intensities (mean per pixel)
        raw_whole_440  = float(np.mean(ch_440[whole_cell]))
        raw_whole_490a = float(np.mean(ch_490a[whole_cell]))
        raw_whole_490b = float(np.mean(ch_490b[whole_cell]))

        # Background-subtracted whole cell
        sub_whole_440  = raw_whole_440  - bg_440
        sub_whole_490a = raw_whole_490a - bg_490a
        sub_whole_490b = raw_whole_490b - bg_490b

        # GP for whole cell
        I_490_whole = (sub_whole_490a + sub_whole_490b) / 2.0
        I_440_whole = sub_whole_440
        denom_whole = I_440_whole + I_490_whole
        gp_whole = (I_440_whole - I_490_whole) / denom_whole if denom_whole != 0 else np.nan

        results.append({
            'Cell_ID'           : int(cell_id),
            'BG_440'            : round(bg_440,  2),
            'BG_490A'           : round(bg_490a, 2),
            'BG_490B'           : round(bg_490b, 2),
            # Membrane measurements
            'Raw_membrane_440'  : round(raw_memb_440,  2),
            'Raw_membrane_490A' : round(raw_memb_490a, 2),
            'Raw_membrane_490B' : round(raw_memb_490b, 2),
            'Sub_membrane_440'  : round(sub_memb_440,  2),
            'Sub_membrane_490A' : round(sub_memb_490a, 2),
            'Sub_membrane_490B' : round(sub_memb_490b, 2),
            'I_490_mean_membrane': round(I_490_memb, 2),
            'GP_membrane'       : round(gp_memb, 4) if not np.isnan(gp_memb) else np.nan,
            # Whole cell measurements
            'Raw_whole_440'     : round(raw_whole_440,  2),
            'Raw_whole_490A'    : round(raw_whole_490a, 2),
            'Raw_whole_490B'    : round(raw_whole_490b, 2),
            'Sub_whole_440'     : round(sub_whole_440,  2),
            'Sub_whole_490A'    : round(sub_whole_490a, 2),
            'Sub_whole_490B'    : round(sub_whole_490b, 2),
            'I_490_mean_whole'  : round(I_490_whole, 2),
            'GP_whole'          : round(gp_whole, 4) if not np.isnan(gp_whole) else np.nan,
        })

    return results


# ── Visualization ─────────────────────────────────────────────────────────────

def save_overview(ch_bright, ch440, labeled_cells, membrane_mask, image_name, out_path):
    """Save an annotated overview PNG with three panels."""
    fig, axes = plt.subplots(1, 3, figsize=(18, 6))

    # Panel 1: Brightfield with auto B&C (2%–98% percentile stretch), cell boundaries + black labels
    vmin_bf = np.percentile(ch_bright, 2)
    vmax_bf = np.percentile(ch_bright, 98)
    axes[0].imshow(ch_bright, cmap='gray', vmin=vmin_bf, vmax=vmax_bf)
    bounds = segmentation.find_boundaries(labeled_cells, mode='outer')
    axes[0].imshow(np.ma.masked_where(~bounds, np.ones_like(bounds)),
                   cmap='autumn', alpha=0.9)
    for prop in measure.regionprops(labeled_cells):
        cy, cx = prop.centroid
        axes[0].text(cx, cy, str(prop.label), color='black',
                     fontsize=9, ha='center', va='center', fontweight='bold',
                     bbox=dict(facecolor='white', alpha=0.55, edgecolor='none', pad=1))
    axes[0].set_title(f'{image_name}\nCell boundaries (brightfield, auto B&C)', fontsize=10)
    axes[0].axis('off')

    # Panel 2: labeled mask
    axes[1].imshow(labeled_cells, cmap='tab20')
    for prop in measure.regionprops(labeled_cells):
        cy, cx = prop.centroid
        axes[1].text(cx, cy, str(prop.label), color='black',
                     fontsize=9, ha='center', va='center', fontweight='bold',
                     bbox=dict(facecolor='white', alpha=0.55, edgecolor='none', pad=1))
    axes[1].set_title('Labeled cell mask', fontsize=10)
    axes[1].axis('off')

    # Panel 3: membrane rings overlaid on 440nm channel
    vmin_440 = np.percentile(ch440, 1)
    vmax_440 = np.percentile(ch440, 99)
    axes[2].imshow(ch440, cmap='gray', vmin=vmin_440, vmax=vmax_440)
    ring_vis = np.ma.masked_where(membrane_mask == 0, membrane_mask)
    axes[2].imshow(ring_vis, cmap='hot', alpha=0.7,
                   vmin=1, vmax=max(labeled_cells.max(), 1))
    for prop in measure.regionprops(labeled_cells):
        cy, cx = prop.centroid
        axes[2].text(cx, cy, str(prop.label), color='black',
                     fontsize=9, ha='center', va='center', fontweight='bold',
                     bbox=dict(facecolor='white', alpha=0.55, edgecolor='none', pad=1))
    axes[2].set_title(f'Membrane rings ({MEMBRANE_THICKNESS_UM}µm) on 440nm', fontsize=10)
    axes[2].axis('off')

    plt.suptitle(image_name, fontsize=12, fontweight='bold')
    plt.tight_layout()
    plt.savefig(out_path, dpi=120, bbox_inches='tight')
    plt.close(fig)


# ── Excel export ──────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="2F5597")
SUBHDR_FILL   = PatternFill("solid", fgColor="4472C4")
IMG_FILL      = PatternFill("solid", fgColor="D6E4F0")
ALT_FILL      = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
IMG_FONT      = Font(bold=True, size=10)
THIN_BORDER   = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'))

COLUMNS = [
    ('Image',               14),
    ('Cell_ID',              8),
    ('BG_440',              10),
    ('BG_490A',             10),
    ('BG_490B',             10),
    # Membrane measurements
    ('Raw_membrane_440',    16),
    ('Raw_membrane_490A',   16),
    ('Raw_membrane_490B',   16),
    ('Sub_membrane_440',    16),
    ('Sub_membrane_490A',   16),
    ('Sub_membrane_490B',   16),
    ('I_490_mean_membrane', 18),
    ('GP_membrane',         12),
    # Whole cell measurements  
    ('Raw_whole_440',       14),
    ('Raw_whole_490A',      14),
    ('Raw_whole_490B',      14),
    ('Sub_whole_440',       14),
    ('Sub_whole_490A',      14),
    ('Sub_whole_490B',      14),
    ('I_490_mean_whole',    16),
    ('GP_whole',            10),
]


def write_excel(all_rows, out_path):
    """Write all results to a nicely formatted Excel file."""
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "GP Results"

    col_names = [c[0] for c in COLUMNS]

    # ── Header row ──
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    # ── Data rows ──
    row_num = 2
    prev_image = None
    img_alt    = False

    for record in all_rows:
        img_name = record['Image']
        if img_name != prev_image:
            img_alt    = not img_alt
            prev_image = img_name

        fill = IMG_FILL if img_alt else ALT_FILL

        for col_idx, (name, _) in enumerate(COLUMNS, start=1):
            val  = record.get(name, '')
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill      = fill
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if name == 'Image':
                cell.font = IMG_FONT
            if name == 'GP' and isinstance(val, float) and not np.isnan(val):
                # colour-code GP: blue=low, red=high, white=zero
                r = max(0, min(255, int((val + 1) / 2 * 255)))
                b = max(0, min(255, int((1 - (val + 1) / 2) * 255)))
                hex_col = f"FF{r:02X}80{b:02X}"
                cell.fill = PatternFill("solid", fgColor=hex_col)

        row_num += 1

    # ── Summary block (one row per image) ──
    ws.cell(row=row_num + 1, column=1, value='SUMMARY (mean ± std per image)').font = Font(bold=True)
    row_num += 2

    summary_header = ['Image', 'N_cells', 'GP_membrane_mean', 'GP_membrane_std', 'GP_membrane_min', 'GP_membrane_max', 'GP_whole_mean', 'GP_whole_std']
    for ci, h in enumerate(summary_header, 1):
        c = ws.cell(row=row_num, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.border = THIN_BORDER; c.alignment = Alignment(horizontal='center')
    row_num += 1

    df = pd.DataFrame(all_rows)
    for img_name, grp in df.groupby('Image'):
        gp_vals_membrane = grp['GP_membrane'].dropna()
        gp_vals_whole = grp['GP_whole'].dropna()
        row_data = [
            img_name, 
            len(gp_vals_membrane),
            round(gp_vals_membrane.mean(), 4) if len(gp_vals_membrane) else np.nan,
            round(gp_vals_membrane.std(),  4) if len(gp_vals_membrane) > 1 else np.nan,
            round(gp_vals_membrane.min(),  4) if len(gp_vals_membrane) else np.nan,
            round(gp_vals_membrane.max(),  4) if len(gp_vals_membrane) else np.nan,
            round(gp_vals_whole.mean(), 4) if len(gp_vals_whole) else np.nan,
            round(gp_vals_whole.std(),  4) if len(gp_vals_whole) > 1 else np.nan,
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=row_num, column=ci, value=val)
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal='center')
        row_num += 1

    wb.save(out_path)
    print(f"  → Excel saved: {out_path}")


# ── Main pipeline ─────────────────────────────────────────────────────────────

def process_czi(filepath, results_dir):
    """Process a single CZI file. Returns list of result dicts."""
    image_name = Path(filepath).stem
    print(f"\n{'─'*60}")
    print(f"Processing: {image_name}")

    # Load channels
    channels, height, width, pixel_size_um = load_czi_channels(filepath)
    print(f"  Image: {height}×{width} px, pixel size: {pixel_size_um:.4f} µm/px")

    required = [CH_440, CH_490_A, CH_490_B]
    for ch in required:
        if ch not in channels:
            raise ValueError(f"  Channel {ch} not found in {filepath}. "
                             f"Available: {sorted(channels.keys())}")

    ch_440  = channels[CH_440].astype(np.float32)
    ch_490a = channels[CH_490_A].astype(np.float32)
    ch_490b = channels[CH_490_B].astype(np.float32)
    ch_bright = channels.get(CH_BRIGHTFIELD, ch_440).astype(np.float32)

    # Segment cells
    print("  Segmenting cells …")
    labeled_cells, n_cells = segment_cells(
        ch_440,
        min_diameter_um=MIN_CELL_DIAMETER_UM,
        max_diameter_um=MAX_CELL_DIAMETER_UM,
        min_circularity=MIN_CIRCULARITY,
        pixel_size_um=pixel_size_um,
        bg_sigma=BG_CORRECTION_SIGMA,
        seg_sigma=SEGMENTATION_SIGMA,
        threshold_percentile=THRESHOLD_PERCENTILE,
        closing_radius=CLOSING_RADIUS,
        watershed_min_dist=WATERSHED_MIN_DIST,
    )
    print(f"  Found {n_cells} cell(s)")

    if n_cells == 0:
        print("  WARNING: No cells found — skipping this image.")
        return []

    # Membrane rings
    print("  Computing membrane rings …")
    membrane_mask = make_membrane_masks(labeled_cells, MEMBRANE_THICKNESS_UM, pixel_size_um)

    # Save masks — each image gets its own subfolder
    img_dir = os.path.join(results_dir, image_name)
    os.makedirs(img_dir, exist_ok=True)

    imwrite(os.path.join(img_dir, f"{image_name}_mask.tif"),
            labeled_cells.astype(np.int32))
    imwrite(os.path.join(img_dir, f"{image_name}_envelope.tif"),
            membrane_mask.astype(np.int32))

    # Overview image
    overview_path = os.path.join(img_dir, f"{image_name}_overview.png")
    save_overview(ch_bright, ch_440, labeled_cells, membrane_mask, image_name, overview_path)
    print(f"  Saved masks + overview → {img_dir}/")

    # Intensity measurements
    print("  Measuring intensities …")
    cell_results = measure_intensities(
        labeled_cells, membrane_mask, ch_440, ch_490a, ch_490b)

    # Attach image name
    rows = []
    for r in cell_results:
        row = {'Image': image_name}
        row.update(r)
        rows.append(row)

    print(f"  GP values: " +
          ", ".join(f"cell {r['Cell_ID']} membrane={r['GP_membrane']:.3f} whole={r['GP_whole']:.3f}" for r in rows))
    return rows


def main():
    script_dir  = os.path.dirname(os.path.abspath(__file__))
    czi_files   = sorted(glob.glob(os.path.join(script_dir, '*.czi')))
    results_dir = os.path.join(script_dir, 'results')
    os.makedirs(results_dir, exist_ok=True)

    if not czi_files:
        print("No .czi files found in the script directory.")
        sys.exit(1)

    print(f"Found {len(czi_files)} CZI file(s): {[Path(f).name for f in czi_files]}")

    all_rows = []
    for filepath in czi_files:
        try:
            rows = process_czi(filepath, results_dir)
            all_rows.extend(rows)
        except Exception as e:
            print(f"  ERROR processing {Path(filepath).name}: {e}")
            import traceback; traceback.print_exc()

    if not all_rows:
        print("\nNo results to save.")
        sys.exit(1)

    # Write Excel
    excel_path = os.path.join(results_dir, 'gp_results.xlsx')
    write_excel(all_rows, excel_path)

    print(f"\n{'═'*60}")
    print(f"Done. Processed {len(czi_files)} image(s), {len(all_rows)} cells total.")
    print(f"Results → {results_dir}/")


if __name__ == '__main__':
    main()
